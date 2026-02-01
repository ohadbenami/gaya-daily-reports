#!/usr/bin/env python3
"""
Daily Deliveries Report - Gaya Foods
=====================================
Sends a daily summary of deliveries to WhatsApp every morning at 07:30.
Groups deliveries by driver and creates a styled Excel file.

Board: הפצה (ID: 5089475109)
Filters by: date4 (ת. הפצה) = today's date
"""

import os
import sys
import json
import requests
from datetime import datetime
from io import BytesIO

# Excel dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

# Configuration
MONDAY_API_URL = "https://api.monday.com/v2"
MONDAY_API_TOKEN = os.environ.get("MONDAY_API_TOKEN")
TIMELINES_API_KEY = os.environ.get("TIMELINES_API_KEY", "f40ecfc9-31e8-4905-a920-b27e5559fabc")
WHATSAPP_PHONE = "972528012869"  # Ohad's number

# Board configuration
BOARD_ID = "5089475109"  # הפצה board

# Column IDs (verified from API)
COLUMNS = {
    "date": "date4",               # ת. הפצה
    "driver": "color_mkz4z0q4",    # נהג (status column)
    "customer": "text_mkz43a4j",   # לקוח
    "city": "text_mkz4zrrm",       # עיר
    "sku": "text_mkz4pcnj",        # מק"ט
    "description": "text_mkz4c904", # תאור מוצר
    "pallets": "numeric_mkz4s8sc", # משטחים
    "order": "text_mkz4n5dn",      # הזמנת לקוח
    "broadcast": "color_mkz4m0mx", # שידור
}

# Driver labels from Monday.com settings (index -> name)
DRIVER_LABELS = {
    0: "שי",
    1: "אורי",
    2: "נאדר",
    3: "שפע תובלה",
    4: "אורי נגלה 2",
    6: "הפצה ביג לוג",
    7: "שי נגלה 2",
    8: "איסוף עצמי",
    9: "סוכן",
    10: "ישראל",
    11: "BL",
}

# Excel styling colors per driver
DRIVER_COLORS = {
    "שי": "C8E6C9",           # Light green
    "אורי": "FFECB3",         # Light amber
    "נאדר": "B3E5FC",         # Light blue
    "שפע תובלה": "E1BEE7",    # Light purple
    "הפצה ביג לוג": "FFCDD2", # Light red
    "איסוף עצמי": "F5F5F5",   # Light gray
    "סוכן": "D7CCC8",         # Light brown
    "ישראל": "B2DFDB",        # Light teal
    "default": "EEEEEE",      # Default gray
}


def log(message):
    """Print with timestamp"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}")


def query_monday_deliveries(target_date: str) -> list:
    """
    Query Monday.com for deliveries on target date.
    Uses items_page_by_column_values to filter by date.

    Args:
        target_date: Date in YYYY-MM-DD format

    Returns:
        List of delivery items
    """
    if not MONDAY_API_TOKEN:
        log("ERROR: MONDAY_API_TOKEN not set")
        return []

    # Query using items_page_by_column_values for date filtering
    query = """
    query ($boardId: ID!, $columnId: String!, $columnValue: String!) {
        items_page_by_column_values(
            board_id: $boardId,
            columns: [{column_id: $columnId, column_values: [$columnValue]}],
            limit: 500
        ) {
            items {
                id
                name
                column_values {
                    id
                    text
                    value
                }
            }
        }
    }
    """

    headers = {
        "Authorization": MONDAY_API_TOKEN,
        "Content-Type": "application/json",
        "API-Version": "2024-10"
    }

    payload = {
        "query": query,
        "variables": {
            "boardId": BOARD_ID,
            "columnId": COLUMNS["date"],
            "columnValue": target_date
        }
    }

    try:
        log(f"Querying Monday.com for deliveries on {target_date}...")
        response = requests.post(MONDAY_API_URL, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()

        if "errors" in data:
            log(f"Monday.com API errors: {data['errors']}")
            # Fallback to regular query
            return query_monday_deliveries_fallback(target_date)

        items = data.get("data", {}).get("items_page_by_column_values", {}).get("items", [])
        log(f"Found {len(items)} deliveries for {target_date}")

        deliveries = []
        for item in items:
            columns = {cv["id"]: cv for cv in item.get("column_values", [])}
            deliveries.append(parse_delivery_item(item, columns))

        return deliveries

    except requests.exceptions.RequestException as e:
        log(f"ERROR: Monday.com API request failed: {e}")
        return query_monday_deliveries_fallback(target_date)
    except Exception as e:
        log(f"ERROR: Unexpected error: {e}")
        return []


def query_monday_deliveries_fallback(target_date: str) -> list:
    """
    Fallback query method - gets all items and filters locally.
    """
    query = """
    query ($boardId: [ID!]!) {
        boards(ids: $boardId) {
            items_page(limit: 500) {
                items {
                    id
                    name
                    column_values {
                        id
                        text
                        value
                    }
                }
            }
        }
    }
    """

    headers = {
        "Authorization": MONDAY_API_TOKEN,
        "Content-Type": "application/json",
        "API-Version": "2024-10"
    }

    payload = {
        "query": query,
        "variables": {"boardId": [BOARD_ID]}
    }

    try:
        log(f"Using fallback query for {target_date}...")
        response = requests.post(MONDAY_API_URL, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()

        items = data.get("data", {}).get("boards", [{}])[0].get("items_page", {}).get("items", [])
        log(f"Retrieved {len(items)} total items")

        deliveries = []
        for item in items:
            columns = {cv["id"]: cv for cv in item.get("column_values", [])}

            # Check date column
            date_col = columns.get(COLUMNS["date"], {})
            date_value = date_col.get("value", "")

            # Parse JSON value to extract date
            if date_value:
                try:
                    date_data = json.loads(date_value)
                    item_date = date_data.get("date", "")
                    if item_date == target_date:
                        deliveries.append(parse_delivery_item(item, columns))
                except (json.JSONDecodeError, TypeError):
                    # Try text match
                    if target_date in str(date_col.get("text", "")):
                        deliveries.append(parse_delivery_item(item, columns))

        log(f"Found {len(deliveries)} deliveries for {target_date}")
        return deliveries

    except Exception as e:
        log(f"ERROR: Fallback query failed: {e}")
        return []


def parse_delivery_item(item: dict, columns: dict) -> dict:
    """Parse a Monday.com item into a delivery dict"""

    # Get driver from status column
    driver_col = columns.get(COLUMNS["driver"], {})
    driver_text = driver_col.get("text", "")
    driver_value = driver_col.get("value", "{}")

    # Parse driver index from JSON value
    driver = driver_text or "לא משויך"
    try:
        if driver_value and driver_value != "{}":
            driver_data = json.loads(driver_value)
            if "index" in driver_data:
                idx = driver_data["index"]
                driver = DRIVER_LABELS.get(idx, driver_text or f"נהג {idx}")
    except (json.JSONDecodeError, TypeError):
        pass

    # Get pallets (numeric)
    pallets_col = columns.get(COLUMNS["pallets"], {})
    pallets_text = pallets_col.get("text", "0")
    try:
        pallets = float(pallets_text) if pallets_text else 0
    except (ValueError, TypeError):
        pallets = 0

    return {
        "id": item.get("id"),
        "name": item.get("name", ""),
        "driver": driver,
        "customer": columns.get(COLUMNS["customer"], {}).get("text", "") or item.get("name", ""),
        "city": columns.get(COLUMNS["city"], {}).get("text", ""),
        "sku": columns.get(COLUMNS["sku"], {}).get("text", ""),
        "description": columns.get(COLUMNS["description"], {}).get("text", ""),
        "pallets": pallets,
        "order": columns.get(COLUMNS["order"], {}).get("text", ""),
    }


def group_by_driver(deliveries: list) -> dict:
    """Group deliveries by driver, sorted alphabetically"""
    grouped = {}
    for d in deliveries:
        driver = d.get("driver", "לא משויך")
        if driver not in grouped:
            grouped[driver] = []
        grouped[driver].append(d)
    return dict(sorted(grouped.items()))


def group_by_driver_and_customer(deliveries: list) -> dict:
    """
    Group deliveries by driver, then by customer.
    Returns: {driver: {customer: [items]}}
    """
    grouped = {}
    for d in deliveries:
        driver = d.get("driver", "לא משויך")
        customer = d.get("customer", "לקוח לא ידוע")

        if driver not in grouped:
            grouped[driver] = {}
        if customer not in grouped[driver]:
            grouped[driver][customer] = []
        grouped[driver][customer].append(d)

    # Sort drivers, then customers within each driver
    return {k: dict(sorted(v.items())) for k, v in sorted(grouped.items())}


def create_excel_report(deliveries: list, target_date: str) -> bytes:
    """
    Create a styled Excel report with TWO sheets:
    1. Summary sheet - only customers with total pallets (main view)
    2. Details sheet - full breakdown of all items

    Returns:
        Excel file as bytes
    """
    wb = openpyxl.Workbook()

    # ============================================
    # SHEET 1: SUMMARY (סיכום) - Main View
    # ============================================
    ws_summary = wb.active
    ws_summary.title = "סיכום"
    ws_summary.sheet_view.rightToLeft = True

    # Styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1D2D44', end_color='1D2D44', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    cell_font = Font(name='Arial', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    date_formatted = datetime.strptime(target_date, "%Y-%m-%d").strftime("%d.%m.%Y")
    ws_summary.merge_cells('A1:D1')
    title_cell = ws_summary['A1']
    title_cell.value = f"הפצות - {date_formatted}"
    title_cell.font = Font(name='Arial', size=20, bold=True, color='1D2D44')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 40

    # Headers for summary - REVERSED for RTL display: משטחים | עיר | לקוח | נהג
    summary_headers = ['משטחים', 'עיר', 'לקוח', 'נהג']
    summary_widths = [14, 22, 45, 18]  # Reversed widths

    for col_idx, (header, width) in enumerate(zip(summary_headers, summary_widths), 1):
        cell = ws_summary.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')  # Bigger font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

    ws_summary.row_dimensions[3].height = 35  # Taller header

    # Group by driver and customer
    grouped = group_by_driver_and_customer(deliveries)

    row = 4
    total_pallets = 0
    total_customers = 0

    for driver, customers in grouped.items():
        driver_color = DRIVER_COLORS.get(driver, DRIVER_COLORS["default"])
        driver_fill = PatternFill(start_color=driver_color, end_color=driver_color, fill_type='solid')

        driver_pallets = 0
        driver_customers = 0

        for customer, items in customers.items():
            total_customers += 1
            driver_customers += 1

            customer_pallets = sum(item.get("pallets", 0) for item in items)
            driver_pallets += customer_pallets
            total_pallets += customer_pallets

            city = items[0].get("city", "") if items else ""

            # Summary row: REVERSED - Pallets | City | Customer | Driver
            row_data = [
                int(customer_pallets) if customer_pallets > 0 else "",
                city,
                customer,
                driver if driver_customers == 1 else ""
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row, column=col_idx, value=value)
                if col_idx == 1:  # Pallets column (now first) - red and bold
                    cell.font = Font(name='Arial', size=14, bold=True, color='E63946')
                else:
                    cell.font = Font(name='Arial', size=12, bold=True)  # Bigger, bold
                cell.fill = driver_fill
                cell.alignment = center_alignment  # ALL CENTERED
                cell.border = thin_border

            ws_summary.row_dimensions[row].height = 32  # Taller rows
            row += 1

        # Driver subtotal - REVERSED columns
        subtotal_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        # Pallets in column A (first)
        pallets_cell = ws_summary.cell(row=row, column=1, value=int(driver_pallets) if driver_pallets > 0 else "")
        pallets_cell.font = Font(name='Arial', size=14, bold=True, color='E63946')
        pallets_cell.fill = subtotal_fill
        pallets_cell.alignment = center_alignment
        pallets_cell.border = thin_border

        # Merge B-D for text
        ws_summary.merge_cells(f'B{row}:D{row}')
        subtotal_cell = ws_summary.cell(row=row, column=2, value=f"סה\"כ {driver}: {driver_customers} לקוחות")
        subtotal_cell.font = Font(name='Arial', size=13, bold=True)
        subtotal_cell.fill = subtotal_fill
        subtotal_cell.alignment = center_alignment
        subtotal_cell.border = thin_border

        ws_summary.row_dimensions[row].height = 35
        row += 2

    # Grand total - REVERSED columns
    row += 1

    # Pallets in column A (first)
    grand_pallets = ws_summary.cell(row=row, column=1, value=int(total_pallets) if total_pallets > 0 else "")
    grand_pallets.font = Font(name='Arial', size=18, bold=True, color='E63946')
    grand_pallets.alignment = center_alignment

    # Merge B-D for text
    ws_summary.merge_cells(f'B{row}:D{row}')
    total_cell = ws_summary.cell(row=row, column=2, value=f"סה\"כ: {total_customers} לקוחות | {len(grouped)} נהגים")
    total_cell.font = Font(name='Arial', size=16, bold=True, color='1D2D44')
    total_cell.alignment = center_alignment

    ws_summary.row_dimensions[row].height = 45

    # ============================================
    # SHEET 2: DETAILS (פירוט) - Full Breakdown
    # ============================================
    ws_details = wb.create_sheet(title="פירוט")
    ws_details.sheet_view.rightToLeft = True

    # Title
    ws_details.merge_cells('A1:G1')
    title_cell = ws_details['A1']
    title_cell.value = f"פירוט הפצות - {date_formatted}"
    title_cell.font = Font(name='Arial', size=18, bold=True, color='1D2D44')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_details.row_dimensions[1].height = 35

    # Headers for details
    detail_headers = ['נהג', 'לקוח', 'עיר', 'מק"ט', 'מוצר', 'כמות', 'הזמנה']
    detail_widths = [12, 28, 14, 12, 40, 8, 12]

    for col_idx, (header, width) in enumerate(zip(detail_headers, detail_widths), 1):
        cell = ws_details.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_details.column_dimensions[get_column_letter(col_idx)].width = width

    ws_details.row_dimensions[3].height = 25

    # Write all detail rows
    row = 4
    for driver, customers in grouped.items():
        driver_color = DRIVER_COLORS.get(driver, DRIVER_COLORS["default"])
        driver_fill = PatternFill(start_color=driver_color, end_color=driver_color, fill_type='solid')

        first_driver_row = True
        for customer, items in customers.items():
            first_customer_row = True
            city = items[0].get("city", "") if items else ""

            for item in items:
                desc = item.get("description", "")
                if len(desc) > 50:
                    desc = desc[:47] + "..."

                row_data = [
                    driver if first_driver_row else "",
                    customer if first_customer_row else "",
                    city if first_customer_row else "",
                    item.get("sku", ""),
                    desc,
                    int(item.get("pallets", 0)) if item.get("pallets", 0) > 0 else "",
                    item.get("order", "")
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_details.cell(row=row, column=col_idx, value=value)
                    cell.font = Font(name='Arial', size=10)
                    cell.fill = driver_fill
                    cell.alignment = center_alignment if col_idx == 6 else center_alignment
                    cell.border = thin_border

                ws_details.row_dimensions[row].height = 20
                row += 1
                first_driver_row = False
                first_customer_row = False

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def format_whatsapp_message(deliveries: list, target_date: str) -> str:
    """Format the WhatsApp summary message"""
    date_formatted = datetime.strptime(target_date, "%Y-%m-%d").strftime("%d.%m.%Y")

    if not deliveries:
        return f"🚚 הפצות היום - {date_formatted}\n\n✅ אין הפצות מתוכננות להיום."

    grouped = group_by_driver(deliveries)

    lines = [f"🚚 הפצות היום - {date_formatted}", ""]

    total_rows = 0
    total_pallets = 0

    for driver, items in grouped.items():
        driver_pallets = sum(item.get("pallets", 0) for item in items)
        driver_rows = len(items)

        total_rows += driver_rows
        total_pallets += driver_pallets

        # Count unique customers
        customers = set(item.get("customer", "") for item in items if item.get("customer"))

        pallets_str = f", {int(driver_pallets)} משטחים" if driver_pallets > 0 else ""
        lines.append(f"📍 {driver} ({len(customers)} לקוחות, {driver_rows} שורות{pallets_str}):")

        # Show unique customers with their cities
        shown_customers = set()
        for item in items:
            customer = item.get("customer", "")
            if customer and customer not in shown_customers and len(shown_customers) < 5:
                city = item.get("city", "")
                city_str = f" - {city}" if city else ""
                # Count pallets for this customer
                cust_pallets = sum(i.get("pallets", 0) for i in items if i.get("customer") == customer)
                pallet_str = f" ({int(cust_pallets)} מש')" if cust_pallets > 0 else ""
                lines.append(f"  • {customer}{city_str}{pallet_str}")
                shown_customers.add(customer)

        if len(customers) > 5:
            lines.append(f"  • ... ועוד {len(customers) - 5} לקוחות")

        lines.append("")

    lines.append("━━━━━━━━━━━━━━━━━━━━━━━━")
    pallets_total = f" | {int(total_pallets)} משטחים" if total_pallets > 0 else ""
    lines.append(f"סה\"כ: {total_rows} שורות | {len(grouped)} נהגים{pallets_total}")

    return "\n".join(lines)


def upload_to_timelines(file_bytes: bytes, filename: str) -> str:
    """Upload file to TimelineAI and return file UID"""
    url = "https://app.timelines.ai/integrations/api/files_upload"
    headers = {"Authorization": f"Bearer {TIMELINES_API_KEY}"}

    files = {
        "file": (filename, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    }

    try:
        log(f"Uploading {filename} to TimelineAI...")
        response = requests.post(url, headers=headers, files=files, timeout=60)
        response.raise_for_status()
        data = response.json()
        file_uid = data.get("uid", "")
        log(f"Upload successful, file UID: {file_uid}")
        return file_uid
    except Exception as e:
        log(f"ERROR: Upload failed: {e}")
        return ""


def send_whatsapp_message(phone: str, text: str = None, file_uid: str = None) -> bool:
    """Send WhatsApp message via TimelineAI"""
    url = "https://app.timelines.ai/integrations/api/messages"
    headers = {
        "Authorization": f"Bearer {TIMELINES_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {"phone": phone}
    if text:
        payload["text"] = text
    if file_uid:
        payload["file_uid"] = file_uid

    try:
        log(f"Sending WhatsApp to {phone}...")
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        log("WhatsApp sent successfully!")
        return True
    except Exception as e:
        log(f"ERROR: WhatsApp send failed: {e}")
        return False


def main():
    """Main execution"""
    log("=" * 60)
    log("🚚 Daily Deliveries Report - Gaya Foods")
    log("=" * 60)

    # Get today's date (or use argument)
    if len(sys.argv) > 1:
        target_date = sys.argv[1]
    else:
        target_date = datetime.now().strftime("%Y-%m-%d")

    log(f"Target date: {target_date}")

    # Query Monday.com
    deliveries = query_monday_deliveries(target_date)

    # Format WhatsApp message
    message = format_whatsapp_message(deliveries, target_date)
    log(f"\nMessage preview:\n{message}\n")

    # Send text message first
    send_whatsapp_message(WHATSAPP_PHONE, text=message)

    # If we have deliveries, create and send Excel
    if deliveries:
        excel_bytes = create_excel_report(deliveries, target_date)
        date_formatted = datetime.strptime(target_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        filename = f"הפצות_{date_formatted.replace('.', '_')}.xlsx"

        # Save locally for artifact
        with open(filename, "wb") as f:
            f.write(excel_bytes)
        log(f"Excel saved locally: {filename}")

        # Upload and send via WhatsApp
        file_uid = upload_to_timelines(excel_bytes, filename)
        if file_uid:
            send_whatsapp_message(WHATSAPP_PHONE, file_uid=file_uid)

    log("=" * 60)
    log("✅ Daily Deliveries Report completed!")
    log("=" * 60)


if __name__ == "__main__":
    main()
