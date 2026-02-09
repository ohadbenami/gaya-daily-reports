#!/usr/bin/env python3
"""
Daily Containers Report - Gaya Foods (Enhanced Version)
- Container list from Monday.com (status, ETA)
- Items from Priority ERP (PORDERITEMS_SUBFORM)
- Landing cost calculations with shipping input cell
- Sends via WhatsApp to Ohad and Kiril
"""

import os
import requests
import json
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Configuration
MONDAY_API_TOKEN = os.environ.get('MONDAY_API_TOKEN')
MONDAY_API_URL = "https://api.monday.com/v2"
PRIORITY_API_HOST = os.environ.get('PRIORITY_API_HOST', 'https://p.priority-connect.online/odata/Priority/tabzfdbb.ini/a230521')
PRIORITY_API_TOKEN = os.environ.get('PRIORITY_API_TOKEN')
PRIORITY_API_PASSWORD = os.environ.get('PRIORITY_API_PASSWORD', 'PAT')
TIMELINES_API_KEY = os.environ.get('TIMELINES_API_KEY')

# Board IDs
ORDERS_BOARD_ID = 1900622333
CURRENCIES_BOARD_ID = 1958318760

# Fixed costs
PORT_COST_ILS = 1107

# Known shipping costs (updated from historical data)
SHIPPING_COSTS = {
    'PO240350': 6643, 'PO240353': 6643,
    'PO250271': 3228, 'PO250272': 3228,
    'PO250299': 2724, 'PO250307': 2724,
    'PO250294': 2724, 'PO250295': 2724, 'PO250296': 2724,
    'PO250344': 2850,
    'PO250345': 3178, 'PO250346': 3178, 'PO250342': 3178,
    'PO250381': 3379,
    'PO250428': 3934,
    'PO250471': 4716, 'PO250472': 4716, 'PO250378': 4716, 'PO250387': 4716,
    'PO250476': 5246, 'PO250477': 5246, 'PO250478': 5246, 'PO250479': 5246, 'PO250483': 5246,
    'PO250379': 3127, 'PO250388': 3127, 'PO250500': 3127,
    'PO250501': 4741,
    'PO250389': 3127, 'PO250502': 3228, 'PO250484': 3228,
}

# Default units per carton by SKU prefix
UNITS_PER_CARTON_DEFAULTS = {
    'TUPP05': 120,  # Small pouches 1/120
}
DEFAULT_UNITS_PER_CARTON = 12

# WhatsApp recipients
RECIPIENTS = [
    {'name': 'אוהד', 'phone': '972528012869'},
    {'name': 'קיריל', 'phone': '972538470070'},
    {'name': 'יובל', 'phone': '972505267110'},
]

# Styles
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SECTION_FILL = PatternFill(start_color="5B768A", end_color="5B768A", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
CALC_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
RED_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

HEADER_FONT = Font(name='Arial', size=14, bold=True, color="FFFFFF")
TITLE_FONT = Font(name='Arial', size=20, bold=True, color="2C3E50")
LABEL_FONT = Font(name='Arial', size=14, bold=True, color="333333")
DATA_FONT = Font(name='Arial', size=14, color="333333")
INPUT_FONT = Font(name='Arial', size=16, bold=True, color="C0392B")
CALC_FONT = Font(name='Arial', size=14, bold=True, color="27AE60")
BIG_NUMBER = Font(name='Arial', size=28, bold=True, color="2C3E50")

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

medium_border = Border(
    left=Side(style='medium', color='333333'),
    right=Side(style='medium', color='333333'),
    top=Side(style='medium', color='333333'),
    bottom=Side(style='medium', color='333333')
)


def monday_query(query):
    """Execute Monday.com GraphQL query"""
    headers = {
        "Authorization": MONDAY_API_TOKEN,
        "Content-Type": "application/json"
    }
    response = requests.post(MONDAY_API_URL, json={"query": query}, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Monday API error: {response.status_code}")
        return None


def priority_query(table, params, retries=3):
    """Execute Priority OData query with retry logic"""
    url = f"{PRIORITY_API_HOST}/{table}"

    for attempt in range(retries):
        try:
            response = requests.get(
                url,
                params=params,
                auth=(PRIORITY_API_TOKEN, PRIORITY_API_PASSWORD),
                headers={'Content-Type': 'application/json'},
                timeout=60
            )
            if response.status_code == 200:
                return response.json().get('value', [])
            elif response.status_code in [502, 503, 504]:
                print(f"Priority API error {response.status_code}, retry {attempt + 1}/{retries}...")
                time.sleep(5 * (attempt + 1))  # Exponential backoff
                continue
            else:
                print(f"Priority API error: {response.status_code} - {response.text}")
                return []
        except requests.exceptions.Timeout:
            print(f"Priority API timeout, retry {attempt + 1}/{retries}...")
            time.sleep(5 * (attempt + 1))
            continue
        except Exception as e:
            print(f"Priority API error: {e}")
            return []

    print("Priority API failed after all retries")
    return []


def fetch_usd_rate():
    """Fetch USD exchange rate from Monday.com"""
    query = f'''
    {{
        boards(ids: [{CURRENCIES_BOARD_ID}]) {{
            items_page(limit: 10) {{
                items {{
                    name
                    column_values {{ id text }}
                }}
            }}
        }}
    }}
    '''
    result = monday_query(query)
    if result and 'data' in result:
        items = result['data']['boards'][0]['items_page']['items']
        for item in items:
            if item['name'] == 'USD':
                for col in item['column_values']:
                    if col['id'] == 'numeric_mkqyfw35':
                        return float(col['text']) if col['text'] else 3.5
    return 3.5


def fetch_containers_from_priority():
    """Fetch Ardo containers with active shipping statuses directly from Priority ERP"""
    # Valid statuses for containers in transit/at port
    valid_statuses = [
        'באוניה',         # On ship - in transit
        'כנ"מ ללא BL',    # At port without BL
    ]

    # Statuses to exclude
    excluded_statuses = [
        'סגור',
        'סגורה',
        'מבוטלת',
        'טיוטא',
        'נשלחה ליצרן',
    ]

    # Build OData filter - Ardo only, exclude closed/cancelled/draft
    exclude_filters = " and ".join([f"STATDES ne '{s}'" for s in excluded_statuses])

    params = {
        '$filter': f"CDES eq 'Ardo Company Ltd' and {exclude_filters}",
        '$select': 'ORDNAME,SUPNAME,CDES,CURDATE,QPRICE,STATDES,IMPFNUM,NOA_ETA,NOA_KONTAINER',
        '$orderby': 'NOA_ETA asc',
        '$top': 100
    }

    data = priority_query('PORDERS', params)

    containers = []
    for order in data:
        status = order.get('STATDES', '')

        # Only include באוניה and כנ"מ ללא BL
        if status not in valid_statuses:
            continue

        po = order.get('ORDNAME', '')
        if not po:
            continue

        # Parse ETA date
        eta = order.get('NOA_ETA', '')
        if eta and 'T' in str(eta):
            eta = str(eta).split('T')[0]

        containers.append({
            'po': po,
            'container': order.get('NOA_KONTAINER', '') or order.get('IMPFNUM', '') or '',
            'supplier': order.get('CDES', '') or order.get('SUPNAME', '') or 'Ardo',
            'eta': eta,
            'fob_total': float(order.get('QPRICE', 0) or 0),
            'currency': '$',
            'status': status,
        })

    print(f"  Statuses found: {set(order.get('STATDES') for order in data)}")
    return containers


def fetch_items_from_priority(po_number):
    """Fetch items for a PO from Priority PORDERITEMS_SUBFORM, with discount applied"""
    params = {
        '$filter': f"ORDNAME eq '{po_number}'",
        '$select': 'ORDNAME,CDES,QPRICE',
        '$expand': 'PORDERITEMS_SUBFORM($select=PARTNAME,PDES,TQUANT,PRICE,QPRICE)'
    }
    data = priority_query('PORDERS', params)

    if not data:
        return []

    items = []
    for order in data:
        all_lines = order.get('PORDERITEMS_SUBFORM', [])

        # Calculate total discount from negative lines (like TUD001)
        gross_total = sum(line.get('QPRICE', 0) for line in all_lines if line.get('QPRICE', 0) > 0)
        total_discount = sum(line.get('QPRICE', 0) for line in all_lines if line.get('QPRICE', 0) < 0)

        # Calculate discount percentage
        discount_pct = abs(total_discount) / gross_total if gross_total > 0 else 0

        # Process only product lines (positive quantity)
        for item in all_lines:
            qty = item.get('TQUANT', 0)
            if qty <= 0:
                continue

            # Apply discount to unit price
            list_price = item.get('PRICE', 0)
            net_price = list_price * (1 - discount_pct)

            sku = item.get('PARTNAME', '')
            items.append({
                'sku': sku,
                'description': item.get('PDES', ''),
                'quantity': int(qty),
                'unit': 'קרט',
                'unit_price': round(net_price, 2),  # Net price after discount
                'units_per_carton': UNITS_PER_CARTON_DEFAULTS.get(sku, DEFAULT_UNITS_PER_CARTON)
            })

    return items


def calculate_days_in_port(eta_str):
    """Calculate days since ETA"""
    if not eta_str:
        return 0
    try:
        eta = datetime.strptime(eta_str, '%Y-%m-%d')
        delta = datetime.now() - eta
        return max(0, delta.days)
    except:
        return 0


def create_summary_sheet(ws, containers, usd_rate):
    """Create summary dashboard sheet with separation between port and on-ship containers"""
    ws.sheet_view.rightToLeft = True

    widths = [3, 15, 18, 20, 14, 14, 12, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Split containers by status
    at_port = [c for c in containers if c['status'] == 'כנ"מ ללא BL']
    on_ship = [c for c in containers if c['status'] == 'באוניה']

    # Sort: at_port by days (descending), on_ship by ETA (ascending)
    at_port = sorted(at_port, key=lambda x: calculate_days_in_port(x['eta']), reverse=True)
    on_ship = sorted(on_ship, key=lambda x: x['eta'] or '9999')

    at_port_fob = sum(c['fob_total'] for c in at_port)
    on_ship_fob = sum(c['fob_total'] for c in on_ship)
    critical_count = sum(1 for c in at_port if calculate_days_in_port(c['eta']) > 30)

    # Title
    ws.merge_cells('B2:H2')
    ws['B2'] = f"🚢 דוח מכולות Ardo - Gaya Foods"
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B3:H3')
    ws['B3'] = f"📅 {datetime.now().strftime('%d.%m.%Y')} | 💵 שער: {usd_rate}"
    ws['B3'].font = Font(size=12, color="7F8C8D")
    ws['B3'].alignment = Alignment(horizontal='center')

    # === KPIs Row 1: Port Containers ===
    row = 5
    PORT_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    SHIP_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

    # At Port count
    ws.merge_cells(f'B{row}:C{row+1}')
    ws[f'B{row}'] = str(len(at_port))
    ws[f'B{row}'].font = Font(name='Arial', size=28, bold=True, color="C0392B")
    ws[f'B{row}'].fill = PORT_FILL
    ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'B{row+2}:C{row+2}')
    ws[f'B{row+2}'] = "⚓ בנמל"
    ws[f'B{row+2}'].font = LABEL_FONT
    ws[f'B{row+2}'].alignment = Alignment(horizontal='center')

    # At Port FOB
    ws.merge_cells(f'D{row}:E{row+1}')
    ws[f'D{row}'] = f"${at_port_fob/1000:.0f}K"
    ws[f'D{row}'].font = Font(name='Arial', size=28, bold=True, color="C0392B")
    ws[f'D{row}'].fill = PORT_FILL
    ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'D{row+2}:E{row+2}')
    ws[f'D{row+2}'] = "FOB בנמל"
    ws[f'D{row+2}'].font = LABEL_FONT
    ws[f'D{row+2}'].alignment = Alignment(horizontal='center')

    # Critical count
    ws.merge_cells(f'F{row}:G{row+1}')
    ws[f'F{row}'] = f"{critical_count} 🔴"
    ws[f'F{row}'].font = Font(name='Arial', size=28, bold=True, color="C0392B")
    ws[f'F{row}'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'F{row+2}:G{row+2}')
    ws[f'F{row+2}'] = "קריטי (>30 יום)"
    ws[f'F{row+2}'].font = LABEL_FONT
    ws[f'F{row+2}'].alignment = Alignment(horizontal='center')

    # === KPIs Row 2: On Ship ===
    row = 9
    # On Ship count
    ws.merge_cells(f'B{row}:C{row+1}')
    ws[f'B{row}'] = str(len(on_ship))
    ws[f'B{row}'].font = Font(name='Arial', size=28, bold=True, color="27AE60")
    ws[f'B{row}'].fill = SHIP_FILL
    ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'B{row+2}:C{row+2}')
    ws[f'B{row+2}'] = "🚢 באוניה"
    ws[f'B{row+2}'].font = LABEL_FONT
    ws[f'B{row+2}'].alignment = Alignment(horizontal='center')

    # On Ship FOB
    ws.merge_cells(f'D{row}:E{row+1}')
    ws[f'D{row}'] = f"${on_ship_fob/1000:.0f}K"
    ws[f'D{row}'].font = Font(name='Arial', size=28, bold=True, color="27AE60")
    ws[f'D{row}'].fill = SHIP_FILL
    ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'D{row+2}:E{row+2}')
    ws[f'D{row+2}'] = "FOB באוניה"
    ws[f'D{row+2}'].font = LABEL_FONT
    ws[f'D{row+2}'].alignment = Alignment(horizontal='center')

    # Total containers
    ws.merge_cells(f'F{row}:G{row+1}')
    ws[f'F{row}'] = str(len(containers))
    ws[f'F{row}'].font = BIG_NUMBER
    ws[f'F{row}'].fill = LIGHT_BLUE
    ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'F{row+2}:G{row+2}')
    ws[f'F{row+2}'] = "סה\"כ מכולות"
    ws[f'F{row+2}'].font = LABEL_FONT
    ws[f'F{row+2}'].alignment = Alignment(horizontal='center')

    # ========== TABLE 1: AT PORT (בנמל) ==========
    row = 14
    ws.merge_cells(f'B{row}:H{row}')
    ws[f'B{row}'] = f"⚓ בנמל - ממתינות לשחרור ({len(at_port)} מכולות | ${at_port_fob:,.0f})"
    ws[f'B{row}'].font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    ws[f'B{row}'].fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    for col in range(2, 9):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        ws.cell(row=row, column=col).border = thin_border

    row += 1
    headers_port = ["#", "הזמנה", "מכולה", "ETA", "FOB $", "ימים בנמל", "גיליון"]
    for col, header in enumerate(headers_port, 2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, cont in enumerate(at_port, 1):
        row += 1
        days = calculate_days_in_port(cont['eta'])

        if days > 30:
            row_fill = RED_FILL
        elif days > 14:
            row_fill = ORANGE_FILL
        else:
            row_fill = GREEN_FILL

        eta_fmt = ''
        if cont['eta']:
            try:
                eta_fmt = datetime.strptime(cont['eta'], '%Y-%m-%d').strftime('%d.%m.%y')
            except:
                eta_fmt = cont['eta']

        values = [i, cont['po'], cont['container'] or '-', eta_fmt or '-',
                  f"${cont['fob_total']:,.0f}", str(days), f"→ {cont['po']}"]

        for col, value in enumerate(values, 2):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if col == 7:  # Days column
                cell.fill = row_fill
                cell.font = Font(name='Arial', size=14, bold=True)

    # ========== TABLE 2: ON SHIP (באוניה) ==========
    row += 2
    ws.merge_cells(f'B{row}:H{row}')
    ws[f'B{row}'] = f"🚢 באוניה - בדרך לישראל ({len(on_ship)} מכולות | ${on_ship_fob:,.0f})"
    ws[f'B{row}'].font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    ws[f'B{row}'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    for col in range(2, 9):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        ws.cell(row=row, column=col).border = thin_border

    row += 1
    headers_ship = ["#", "הזמנה", "מכולה", "ETA צפוי", "FOB $", "ימים להגעה", "גיליון"]
    for col, header in enumerate(headers_ship, 2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, cont in enumerate(on_ship, 1):
        row += 1

        # Calculate days until arrival
        days_until = 0
        eta_fmt = ''
        if cont['eta']:
            try:
                eta_date = datetime.strptime(cont['eta'], '%Y-%m-%d')
                eta_fmt = eta_date.strftime('%d.%m.%y')
                days_until = max(0, (eta_date - datetime.now()).days)
            except:
                eta_fmt = cont['eta']

        values = [i, cont['po'], cont['container'] or '-', eta_fmt or '-',
                  f"${cont['fob_total']:,.0f}", str(days_until), f"→ {cont['po']}"]

        for col, value in enumerate(values, 2):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if col == 7:  # Days column
                cell.fill = SHIP_FILL


def create_container_sheet(wb, container, items, usd_rate):
    """Create detailed sheet for a container with items from Priority"""
    sheet_name = container['po'][:31]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.rightToLeft = True
    
    widths = [4, 20, 12, 40, 10, 14, 14, 14, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Title
    row = 2
    ws.merge_cells(f'B{row}:J{row}')
    ws[f'B{row}'] = f"📦 עלות נחיתה - מכולה {container['po']}"
    ws[f'B{row}'].font = TITLE_FONT
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    
    # Container details
    row = 4
    eta_fmt = ''
    if container['eta']:
        try:
            eta_fmt = datetime.strptime(container['eta'], '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            eta_fmt = container['eta']
    
    ws.cell(row=row, column=2, value="מספר מכולה:").font = LABEL_FONT
    ws.cell(row=row, column=3, value=container['container'] or '-').font = DATA_FONT
    ws.cell(row=row, column=5, value="ספק:").font = LABEL_FONT
    ws.cell(row=row, column=6, value=container['supplier']).font = DATA_FONT
    row += 1
    ws.cell(row=row, column=2, value="ETA:").font = LABEL_FONT
    ws.cell(row=row, column=3, value=eta_fmt or '-').font = DATA_FONT
    ws.cell(row=row, column=5, value="סטטוס:").font = LABEL_FONT
    ws.cell(row=row, column=6, value=container['status']).font = DATA_FONT
    
    # Parameters section
    row += 2
    ws.merge_cells(f'B{row}:J{row}')
    ws[f'B{row}'] = "⚙️ פרמטרים לחישוב"
    ws[f'B{row}'].font = HEADER_FONT
    ws[f'B{row}'].fill = SECTION_FILL
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    for col in range(2, 11):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = thin_border
    
    row += 1
    rate_row = row
    ws.cell(row=row, column=2, value="שער דולר:").font = LABEL_FONT
    rate_cell = ws.cell(row=row, column=3, value=usd_rate)
    rate_cell.font = DATA_FONT
    rate_cell.number_format = '0.000'
    rate_cell.fill = LIGHT_BLUE
    rate_cell.border = thin_border
    
    ws.cell(row=row, column=5, value="עלות נמל (₪):").font = LABEL_FONT
    ws.cell(row=row, column=6, value=PORT_COST_ILS).font = DATA_FONT
    ws.cell(row=row, column=6).fill = LIGHT_BLUE
    ws.cell(row=row, column=6).border = thin_border
    
    ws.cell(row=row, column=8, value="מכס:").font = LABEL_FONT
    ws.cell(row=row, column=9, value="0%").font = DATA_FONT
    
    # Shipping input - pre-fill from known costs
    row += 1
    shipping_row = row
    known_shipping = SHIPPING_COSTS.get(container['po'], 0)
    ws.cell(row=row, column=2, value="👇 עלות הובלה סה\"כ ($):").font = Font(name='Arial', size=14, bold=True, color="C0392B")
    shipping_cell = ws.cell(row=row, column=3, value=known_shipping if known_shipping > 0 else "")
    shipping_cell.fill = INPUT_FILL
    shipping_cell.border = medium_border
    shipping_cell.font = INPUT_FONT
    shipping_cell.number_format = '$#,##0'
    shipping_cell.alignment = Alignment(horizontal='center')
    if known_shipping > 0:
        ws.cell(row=row, column=5, value="✅ מעודכן").font = Font(size=12, italic=True, color="27AE60")
    else:
        ws.cell(row=row, column=5, value="← מלא כאן").font = Font(size=12, italic=True, color="888888")
    
    # Items table
    row += 2
    ws.merge_cells(f'B{row}:J{row}')
    ws[f'B{row}'] = "📋 פירוט מק\"טים ועלות נחיתה"
    ws[f'B{row}'].font = HEADER_FONT
    ws[f'B{row}'].fill = SECTION_FILL
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    for col in range(2, 11):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = thin_border
    
    row += 1
    headers = ["#", "מק\"ט", "כמות", "תיאור", "FOB/יח' $", "FOB סה\"כ $", "הובלה/יח' $", "נמל/יח' ₪", "עלות נחיתה/יח' ₪"]
    for col, header in enumerate(headers, 2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row].height = 35
    
    if not items:
        row += 1
        ws.merge_cells(f'B{row}:J{row}')
        ws[f'B{row}'] = "אין פריטים להציג"
        ws[f'B{row}'].font = Font(size=12, italic=True, color="888888")
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        return
    
    total_units = sum(item['quantity'] for item in items if item['quantity'] > 0)
    if total_units == 0:
        total_units = 1
    
    row += 1
    first_data_row = row
    shipping_ref = f"$C${shipping_row}"
    rate_ref = f"$C${rate_row}"
    
    for i, item in enumerate(items, 1):
        if item['quantity'] <= 0:
            continue
        
        fob_total = item['quantity'] * item['unit_price']
        
        ws.cell(row=row, column=2, value=i).font = DATA_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=3, value=item['sku']).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=4, value=item['quantity']).font = DATA_FONT
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        
        desc = item['description'][:35] if item['description'] else ''
        ws.cell(row=row, column=5, value=desc).font = DATA_FONT
        
        ws.cell(row=row, column=6, value=item['unit_price']).font = DATA_FONT
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=7, value=fob_total).font = DATA_FONT
        ws.cell(row=row, column=7).number_format = '$#,##0'
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
        
        # Shipping per unit formula
        ship_formula = f'=IF({shipping_ref}="","",{shipping_ref}/{total_units})'
        ws.cell(row=row, column=8, value=ship_formula).font = CALC_FONT
        ws.cell(row=row, column=8).fill = CALC_FILL
        ws.cell(row=row, column=8).number_format = '$#,##0.00'
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
        
        # Port per unit
        port_per_unit = PORT_COST_ILS / total_units
        ws.cell(row=row, column=9, value=port_per_unit).font = DATA_FONT
        ws.cell(row=row, column=9).fill = CALC_FILL
        ws.cell(row=row, column=9).number_format = '₪#,##0.00'
        ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
        
        # Landing cost formula
        landing_formula = f'=IF(H{row}="",F{row}*{rate_ref}+I{row},(F{row}+H{row})*{rate_ref}+I{row})'
        ws.cell(row=row, column=10, value=landing_formula)
        ws.cell(row=row, column=10).font = Font(name='Arial', size=14, bold=True, color="27AE60")
        ws.cell(row=row, column=10).fill = CALC_FILL
        ws.cell(row=row, column=10).number_format = '₪#,##0.00'
        ws.cell(row=row, column=10).alignment = Alignment(horizontal='center')
        
        for col in range(2, 11):
            ws.cell(row=row, column=col).border = thin_border
        ws.row_dimensions[row].height = 25
        row += 1
    
    last_data_row = row - 1
    
    # Totals
    ws.merge_cells(f'B{row}:E{row}')
    ws.cell(row=row, column=2, value="סה\"כ").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    for col in range(2, 6):
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).border = thin_border
    
    ws.cell(row=row, column=6).fill = HEADER_FILL
    ws.cell(row=row, column=6).border = thin_border
    
    ws.cell(row=row, column=7, value=f"=SUM(G{first_data_row}:G{last_data_row})").font = HEADER_FONT
    ws.cell(row=row, column=7).fill = HEADER_FILL
    ws.cell(row=row, column=7).number_format = '$#,##0'
    ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=7).border = thin_border
    
    for col in [8, 9, 10]:
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).border = thin_border
    
    # Legend
    row += 2
    ws.merge_cells(f'B{row}:J{row}')
    ws[f'B{row}'] = f"📝 מלא הובלה בתא הצהוב ← החישובים יתעדכנו | סה\"כ יחידות: {total_units:,}"
    ws[f'B{row}'].font = Font(size=11, italic=True, color="666666")
    ws[f'B{row}'].alignment = Alignment(horizontal='center')


def create_all_items_sheet(wb, containers_with_items, usd_rate):
    """Create consolidated sheet with all items from all containers, sorted by ETA"""
    ws = wb.create_sheet(title="כל המק\"טים", index=1)  # Insert after summary
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = "FF0000"  # RED tab

    # Column widths - 14 columns now
    widths = [4, 12, 14, 14, 8, 8, 35, 12, 14, 14, 12, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells('A2:N2')
    ws['A2'] = "📦 כל המק\"טים - ממוין לפי תאריך הגעה"
    ws['A2'].font = TITLE_FONT
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:N3')
    ws['A3'] = f"📅 {datetime.now().strftime('%d.%m.%Y')} | 💵 שער: {usd_rate} | עלויות הובלה מעודכנות"
    ws['A3'].font = Font(size=12, color="7F8C8D")
    ws['A3'].alignment = Alignment(horizontal='center')

    # Flatten and sort all items by ETA
    all_items = []
    for container, items in containers_with_items:
        total_units_in_container = sum(item['quantity'] for item in items if item['quantity'] > 0)
        shipping_cost = SHIPPING_COSTS.get(container['po'], 0)
        for item in items:
            if item['quantity'] <= 0:
                continue
            units_per_carton = item.get('units_per_carton', DEFAULT_UNITS_PER_CARTON)
            all_items.append({
                'po': container['po'],
                'eta': container['eta'],
                'status': container['status'],
                'container_fob': container['fob_total'],
                'total_units_in_container': total_units_in_container,
                'shipping_cost': shipping_cost,
                'units_per_carton': units_per_carton,
                **item
            })

    # Sort by ETA (earliest first), then by PO
    all_items.sort(key=lambda x: (x['eta'] or '9999', x['po']))

    # Headers row - 14 columns with units per carton
    row = 5
    headers = ["#", "מק\"ט", "מספר PO", "תאריך כניסה", "כמות", "יח'/קרט", "תיאור",
               "FOB/יח' $", "FOB סה\"כ $", "עלות הובלה",
               "הובלה/יח' $", "נמל/יח' ₪", "עלות נחיתה/קרט ₪", "עלות סופית ליחידה"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row].height = 45

    # Track which POs we've already added shipping input for
    po_shipping_cells = {}

    row += 1
    first_data_row = row

    for i, item in enumerate(all_items, 1):
        fob_per_unit = item['unit_price']
        fob_total = item['quantity'] * fob_per_unit
        total_units = item['total_units_in_container'] or 1
        shipping_cost = item.get('shipping_cost', 0)
        units_per_carton = item.get('units_per_carton', DEFAULT_UNITS_PER_CARTON)

        # Format ETA
        eta_fmt = ''
        if item['eta']:
            try:
                eta_fmt = datetime.strptime(item['eta'], '%Y-%m-%d').strftime('%d.%m.%y')
            except:
                eta_fmt = item['eta']

        # Row fill based on status
        if item['status'] == 'כנ"מ ללא BL':
            row_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

        # Column 1: #
        ws.cell(row=row, column=1, value=i).font = DATA_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=1).border = thin_border

        # Column 2: SKU
        ws.cell(row=row, column=2, value=item['sku']).font = Font(name='Arial', size=14, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).border = thin_border

        # Column 3: PO
        ws.cell(row=row, column=3, value=item['po']).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).border = thin_border

        # Column 4: ETA
        cell = ws.cell(row=row, column=4, value=eta_fmt or '-')
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = row_fill

        # Column 5: Quantity
        ws.cell(row=row, column=5, value=item['quantity']).font = DATA_FONT
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5).border = thin_border

        # Column 6: Units per carton
        ws.cell(row=row, column=6, value=units_per_carton).font = DATA_FONT
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6).border = thin_border

        # Column 7: Description
        desc = item['description'][:35] if item['description'] else ''
        ws.cell(row=row, column=7, value=desc).font = DATA_FONT
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=7).border = thin_border

        # Column 8: FOB per carton
        ws.cell(row=row, column=8, value=fob_per_unit).font = DATA_FONT
        ws.cell(row=row, column=8).number_format = '$#,##0.00'
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=8).border = thin_border

        # Column 9: FOB total
        ws.cell(row=row, column=9, value=fob_total).font = DATA_FONT
        ws.cell(row=row, column=9).number_format = '$#,##0'
        ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=9).border = thin_border

        # Column 10: Shipping cost (yellow input, pre-filled if known)
        if item['po'] not in po_shipping_cells:
            cell = ws.cell(row=row, column=10, value=shipping_cost if shipping_cost > 0 else "")
            cell.fill = INPUT_FILL
            cell.border = medium_border
            cell.font = INPUT_FONT
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal='center')
            po_shipping_cells[item['po']] = f"$J${row}"
        else:
            cell = ws.cell(row=row, column=10, value=f"={po_shipping_cells[item['po']]}")
            cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
            cell.border = thin_border
            cell.font = DATA_FONT
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal='center')

        shipping_ref = po_shipping_cells[item['po']]

        # Column 11: Shipping per carton
        ship_formula = f'=IF({shipping_ref}="","",{shipping_ref}/{total_units})'
        cell = ws.cell(row=row, column=11, value=ship_formula)
        cell.font = CALC_FONT
        cell.fill = CALC_FILL
        cell.number_format = '$#,##0.00'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

        # Column 12: Port per carton
        port_per_unit = PORT_COST_ILS / total_units
        cell = ws.cell(row=row, column=12, value=port_per_unit)
        cell.font = DATA_FONT
        cell.fill = CALC_FILL
        cell.number_format = '₪#,##0.00'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

        # Column 13: Landing cost per CARTON
        landing_formula = f'=IF(K{row}="",H{row}*{usd_rate}+L{row},(H{row}+K{row})*{usd_rate}+L{row})'
        cell = ws.cell(row=row, column=13, value=landing_formula)
        cell.font = CALC_FONT
        cell.fill = CALC_FILL
        cell.number_format = '₪#,##0.00'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

        # Column 14: Final cost per UNIT (divide by units per carton)
        final_formula = f'=M{row}/F{row}'
        cell = ws.cell(row=row, column=14, value=final_formula)
        cell.font = Font(name='Arial', size=14, bold=True, color="27AE60")
        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        cell.number_format = '₪#,##0.00'
        cell.alignment = Alignment(horizontal='center')
        cell.border = medium_border

        ws.row_dimensions[row].height = 28
        row += 1

    last_data_row = row - 1

    # Totals row
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="סה\"כ").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).border = thin_border

    # Sum of quantities
    cell = ws.cell(row=row, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

    for col in [6, 7, 8]:
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).border = thin_border

    # Sum of FOB total
    cell = ws.cell(row=row, column=9, value=f"=SUM(I{first_data_row}:I{last_data_row})")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.number_format = '$#,##0'
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

    for col in [10, 11, 12, 13, 14]:
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).border = thin_border

    # Legend
    row += 2
    ws.merge_cells(f'A{row}:N{row}')
    ws[f'A{row}'] = "📝 עלות סופית ליחידה = עלות נחיתה לקרטון ÷ יחידות בקרטון | 🟠 כתום = בנמל | 🟢 ירוק = באוניה"
    ws[f'A{row}'].font = Font(size=11, italic=True, color="666666")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'A{row}'].font = Font(size=11, italic=True, color="666666")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')


def create_excel_report(containers, usd_rate):
    """Create full Excel with summary + all items sheet + per-container sheets"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "סיכום מכולות"

    create_summary_sheet(ws, containers, usd_rate)

    # Collect all containers with their items
    containers_with_items = []
    for container in containers:
        print(f"  Fetching items for {container['po']}...")
        items = fetch_items_from_priority(container['po'])
        print(f"    Found {len(items)} items")
        containers_with_items.append((container, items))

    # Create consolidated "all items" sheet (after summary)
    print("  Creating consolidated items sheet...")
    create_all_items_sheet(wb, containers_with_items, usd_rate)

    # Create individual container sheets
    for container, items in containers_with_items:
        create_container_sheet(wb, container, items, usd_rate)

    filename = f"דוח_מכולות_כנמ_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)
    return filename


def upload_file(filepath, retries=3):
    """Upload file to TimelineAI with retry"""
    url = "https://app.timelines.ai/integrations/api/files_upload"
    headers = {"Authorization": f"Bearer {TIMELINES_API_KEY}"}

    for attempt in range(retries):
        try:
            with open(filepath, 'rb') as f:
                response = requests.post(url, headers=headers, files={'file': f}, timeout=30)
            if response.status_code == 200:
                return response.json().get('data', {}).get('uid')
            print(f"Upload error: {response.status_code}, retry {attempt + 1}/{retries}...")
        except Exception as e:
            print(f"Upload exception: {e}, retry {attempt + 1}/{retries}...")
        time.sleep(3 * (attempt + 1))

    print("Upload failed after all retries")
    return None


def send_whatsapp(phone, file_uid, text, retries=2):
    """Send WhatsApp message with file and retry"""
    url = "https://app.timelines.ai/integrations/api/messages"
    headers = {"Authorization": f"Bearer {TIMELINES_API_KEY}", "Content-Type": "application/json"}

    for attempt in range(retries):
        try:
            response = requests.post(url, headers=headers,
                                     json={"phone": phone, "file_uid": file_uid, "text": text},
                                     timeout=30)
            if response.status_code == 200:
                return True
            print(f"WhatsApp send error {response.status_code} for {phone}, retry {attempt + 1}...")
        except Exception as e:
            print(f"WhatsApp exception for {phone}: {e}, retry {attempt + 1}...")
        time.sleep(3)

    return False


def main():
    print("💵 Fetching USD rate...")
    usd_rate = fetch_usd_rate()
    print(f"USD Rate: {usd_rate}")

    print("🚢 Fetching Ardo containers from Priority...")
    containers = fetch_containers_from_priority()

    if not containers:
        print("No containers found")
        return

    print(f"Found {len(containers)} containers")

    # Check shipping cost coverage
    with_shipping = [c for c in containers if SHIPPING_COSTS.get(c['po'], 0) > 0]
    without_shipping = [c for c in containers if SHIPPING_COSTS.get(c['po'], 0) == 0]
    print(f"  Shipping costs: {len(with_shipping)}/{len(containers)} ({len(without_shipping)} missing)")
    if without_shipping:
        print(f"  ⚠️ Missing: {', '.join(c['po'] for c in without_shipping)}")

    print("📊 Creating Excel with items from Priority...")
    filename = create_excel_report(containers, usd_rate)
    print(f"Created: {filename}")

    print("📤 Uploading...")
    file_uid = upload_file(filename)
    if not file_uid:
        print("Upload failed")
        return

    print("📱 Sending WhatsApp...")
    today = datetime.now().strftime('%d.%m.%Y')
    at_port = [c for c in containers if c['status'] == 'כנ"מ ללא BL']
    on_ship = [c for c in containers if c['status'] == 'באוניה']
    critical = sum(1 for c in at_port if calculate_days_in_port(c['eta']) > 30)
    at_port_fob = sum(c['fob_total'] for c in at_port)
    on_ship_fob = sum(c['fob_total'] for c in on_ship)

    text = f"🚢 דוח מכולות Ardo - {today}\n"
    text += f"⚓ {len(at_port)} בנמל (${at_port_fob/1000:.0f}K)\n"
    text += f"🚢 {len(on_ship)} באוניה (${on_ship_fob/1000:.0f}K)\n"
    text += f"📦 סה\"כ: {len(containers)} מכולות"
    if critical > 0:
        text += f"\n🔴 {critical} קריטי (>30 יום בנמל!)"
    text += f"\n💰 הובלה: {len(with_shipping)}/{len(containers)} מעודכנים"
    if without_shipping:
        text += f"\n⚠️ חסר הובלה: {', '.join(c['po'] for c in without_shipping)}"
    text += "\n\n🤖 עדכון יומי אוטומטי"

    for r in RECIPIENTS:
        ok = send_whatsapp(r['phone'], file_uid, text)
        print(f"{'✅' if ok else '❌'} {r['name']}: {r['phone']}")

    print("\n✅ Done!")


if __name__ == "__main__":
    main()
