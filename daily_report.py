#!/usr/bin/env python3
"""
Daily Uninvoiced Delivery Notes Report - Gaya Foods
Queries Priority ERP, creates styled Excel, sends via WhatsApp.
"""

import os
import json
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile

# --- Configuration from environment variables ---
PRIORITY_API_USER = os.environ["PRIORITY_API_USER"]
PRIORITY_API_PASS = os.environ["PRIORITY_API_PASS"]
TIMELINEAI_TOKEN = os.environ["TIMELINEAI_TOKEN"]
WHATSAPP_PHONE = os.environ.get("WHATSAPP_PHONE", "972528012869")

PRIORITY_BASE_URL = "https://p.priority-connect.online/odata/Priority/tabzfdbb.ini/a230521"


def query_priority():
    """Query Priority ERP for uninvoiced delivery notes (last 3 months)."""
    three_months_ago = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%dT00:00:00+02:00")

    url = f"{PRIORITY_BASE_URL}/DOCUMENTS_D"
    params = {
        "$filter": f"IVALL eq 'N' and CURDATE ge {three_months_ago} and QPRICE gt 0",
        "$expand": "TRANSORDER_D_SUBFORM($select=PARTNAME,TQUANT)",
        "$orderby": "CURDATE desc"
    }

    response = requests.get(
        url,
        params=params,
        auth=(PRIORITY_API_USER, PRIORITY_API_PASS),
        headers={"Accept": "application/json"},
        timeout=60
    )
    response.raise_for_status()
    return response.json().get("value", [])


def process_documents(raw_docs):
    """Process raw documents into structured data with pallet counts."""
    documents = []
    for doc in raw_docs:
        lines = doc.get("TRANSORDER_D_SUBFORM", [])
        num_pallets = len(lines)
        total_qty = sum(line.get("TQUANT", 0) for line in lines)

        documents.append({
            "DOCNO": doc.get("DOCNO", ""),
            "CDES": doc.get("CDES", ""),
            "CURDATE": doc.get("CURDATE", "")[:10],
            "QPRICE": doc.get("QPRICE", 0),
            "STATDES": doc.get("STATDES", ""),
            "QTY": int(total_qty),
            "PALLETS": num_pallets
        })

    return documents


def create_excel(documents):
    """Create a styled Excel file from the documents data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "תעודות משלוח פתוחות"
    ws.sheet_view.rightToLeft = True

    # Styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    number_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    status_colors = {
        "ת. לאוברסיז": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "סופית": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "ממתין לחן": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }

    # Title row
    today_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"תעודות משלוח ללא חשבונית - {today_str}"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ["#", "מס' תעודה", "לקוח", "תאריך", "סכום ₪", "קרטונים", "משטחים", "סטטוס"]
    col_widths = [5, 14, 30, 12, 14, 12, 12, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    total_price = 0
    total_qty = 0
    total_pallets = 0

    for idx, doc in enumerate(documents, 1):
        row = idx + 3

        # Format date
        date_str = doc["CURDATE"]
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            date_formatted = date_obj.strftime("%d.%m.%Y")
        except:
            date_formatted = date_str

        values = [
            idx,
            doc["DOCNO"],
            doc["CDES"],
            date_formatted,
            doc["QPRICE"],
            doc["QTY"],
            doc["PALLETS"],
            doc["STATDES"]
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # Number formatting
            if col_idx == 5:  # Price
                cell.number_format = '#,##0'

            # Status color
            if col_idx == 8 and value in status_colors:
                cell.fill = status_colors[value]

        # Alternate row coloring
        if idx % 2 == 0:
            for col_idx in range(1, 9):
                if ws.cell(row=row, column=col_idx).fill == PatternFill():
                    ws.cell(row=row, column=col_idx).fill = PatternFill(
                        start_color="F2F7FC", end_color="F2F7FC", fill_type="solid"
                    )

        total_price += doc["QPRICE"]
        total_qty += doc["QTY"]
        total_pallets += doc["PALLETS"]

    # Summary row
    summary_row = len(documents) + 4
    summary_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    summary_font = Font(name="Arial", size=11, bold=True)

    ws.cell(row=summary_row, column=1, value="").fill = summary_fill
    ws.cell(row=summary_row, column=2, value="סה\"כ").font = summary_font
    ws.cell(row=summary_row, column=2).fill = summary_fill
    ws.cell(row=summary_row, column=2).alignment = data_alignment
    ws.cell(row=summary_row, column=3, value=f"{len(documents)} תעודות").font = summary_font
    ws.cell(row=summary_row, column=3).fill = summary_fill
    ws.cell(row=summary_row, column=3).alignment = data_alignment
    ws.cell(row=summary_row, column=4, value="").fill = summary_fill

    price_cell = ws.cell(row=summary_row, column=5, value=total_price)
    price_cell.font = summary_font
    price_cell.fill = summary_fill
    price_cell.alignment = data_alignment
    price_cell.number_format = '#,##0'

    qty_cell = ws.cell(row=summary_row, column=6, value=total_qty)
    qty_cell.font = summary_font
    qty_cell.fill = summary_fill
    qty_cell.alignment = data_alignment

    pallets_cell = ws.cell(row=summary_row, column=7, value=total_pallets)
    pallets_cell.font = summary_font
    pallets_cell.fill = summary_fill
    pallets_cell.alignment = data_alignment

    ws.cell(row=summary_row, column=8, value="").fill = summary_fill

    for col_idx in range(1, 9):
        ws.cell(row=summary_row, column=col_idx).border = thin_border

    # Save to temp file
    filepath = os.path.join(tempfile.gettempdir(), f"תעודות_משלוח_פתוחות_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    wb.save(filepath)
    return filepath


def upload_to_timelineai(filepath):
    """Upload file to TimelineAI and return the file UID."""
    url = "https://app.timelines.ai/integrations/api/files_upload"
    headers = {"Authorization": f"Bearer {TIMELINEAI_TOKEN}"}

    with open(filepath, "rb") as f:
        files = {"file": (os.path.basename(filepath), f)}
        response = requests.post(url, headers=headers, files=files, timeout=30)

    response.raise_for_status()
    data = response.json()
    return data["data"]["uid"]


def send_whatsapp(file_uid, doc_count, total_price, total_pallets):
    """Send the file via WhatsApp."""
    url = "https://app.timelines.ai/integrations/api/messages"
    headers = {
        "Authorization": f"Bearer {TIMELINEAI_TOKEN}",
        "Content-Type": "application/json"
    }

    today_str = datetime.now().strftime("%d.%m.%Y")
    text = (
        f"תעודות משלוח פתוחות - {today_str}\n"
        f"{doc_count} תעודות | {total_pallets} משטחים | {total_price:,.0f} ₪"
    )

    payload = {
        "phone": WHATSAPP_PHONE,
        "text": text,
        "file_uid": file_uid
    }

    response = requests.post(url, headers=headers, json=payload, timeout=30)
    response.raise_for_status()
    return response.json()


def main():
    print(f"[{datetime.now()}] Starting daily report...")

    # 1. Query Priority
    print("Querying Priority ERP...")
    raw_docs = query_priority()
    print(f"  Found {len(raw_docs)} uninvoiced delivery notes")

    if not raw_docs:
        print("No uninvoiced delivery notes found. Exiting.")
        return

    # 2. Process documents
    documents = process_documents(raw_docs)
    total_price = sum(d["QPRICE"] for d in documents)
    total_pallets = sum(d["PALLETS"] for d in documents)

    print(f"  Total: {total_price:,.0f} ILS, {total_pallets} pallets")

    # 3. Create Excel
    print("Creating Excel...")
    filepath = create_excel(documents)
    print(f"  Saved: {filepath}")

    # 4. Upload to TimelineAI
    print("Uploading to TimelineAI...")
    file_uid = upload_to_timelineai(filepath)
    print(f"  File UID: {file_uid}")

    # 5. Send via WhatsApp
    print("Sending via WhatsApp...")
    result = send_whatsapp(file_uid, len(documents), total_price, total_pallets)
    print(f"  Sent! Message UID: {result.get('data', {}).get('message_uid', 'N/A')}")

    # Cleanup
    os.remove(filepath)
    print(f"[{datetime.now()}] Done!")


if __name__ == "__main__":
    main()
